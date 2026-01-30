import os
import re
import sys
from collections import defaultdict
from typing import Dict, List, Optional, Tuple

from bs4 import BeautifulSoup
from openpyxl import load_workbook

DOCS_DIR = os.path.join(os.path.dirname(__file__), "..", "docs")
APP_DIR  = os.path.join(os.path.dirname(__file__), "..", "app")

ROSTERS_XLSX = os.path.join(APP_DIR, "rosters.xlsx")
TEAM_NAMES_XLSX = os.path.join(DOCS_DIR, "Team_Names.xlsx")

OUT_PLAYER = os.path.join(DOCS_DIR, "Player_Pooh_Summary.html")
OUT_BY_TEAM = os.path.join(DOCS_DIR, "Pooh_Summary_By_Team.html")

# ----------------------------
# Helpers
# ----------------------------
def parse_cap_pd(argv) -> Optional[int]:
    # optional arg: PD7
    if len(argv) < 2:
        return None
    s = argv[1].strip().upper()
    m = re.fullmatch(r"PD(\d+)", s)
    if not m:
        raise SystemExit("Usage: python app/build_player_pooh_summary.py [PD7]")
    return int(m.group(1))

def pd_num_from_filename(fn: str) -> Optional[int]:
    m = re.search(r"Final_Players_PD(\d+)\.html$", fn)
    return int(m.group(1)) if m else None

def norm_name(name: str) -> str:
    s = (name or "").lower()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\b(jr|sr|ii|iii|iv)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def safe_int(x) -> int:
    try:
        return int(str(x).strip())
    except:
        return 0

def safe_float(x) -> float:
    try:
        return float(str(x).strip())
    except:
        return 0.0

def html_read_table(path: str) -> Tuple[List[str], List[List[str]]]:
    with open(path, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f.read(), "html.parser")

    table = soup.find("table")
    if not table:
        return [], []

    headers = [th.get_text(strip=True) for th in table.find_all("th")]
    headers_l = [h.lower() for h in headers]

    rows = []
    for tr in table.find_all("tr")[1:]:
        tds = tr.find_all("td")
        if not tds:
            continue
        rows.append([td.get_text(strip=True) for td in tds])

    return headers_l, rows

def idx(headers_l: List[str], *cands: str) -> Optional[int]:
    for c in cands:
        c = c.lower()
        if c in headers_l:
            return headers_l.index(c)
    return None

# ----------------------------
# Team Names Map (Owner -> Team Name)
# ----------------------------
def load_team_name_map() -> Dict[str, str]:
    """
    docs/Team_Names.xlsx:
      headers: Owner, Team Name
    """
    if not os.path.exists(TEAM_NAMES_XLSX):
        print(f"NOTE: Missing {TEAM_NAMES_XLSX}. Using names as-is.")
        return {}

    wb = load_workbook(TEAM_NAMES_XLSX, data_only=True)
    ws = wb.active

    headers = [("" if c.value is None else str(c.value).strip()) for c in ws[1]]
    headers_l = [h.lower() for h in headers]

    def col(name: str) -> Optional[int]:
        n = name.lower()
        if n in headers_l:
            return headers_l.index(n) + 1
        return None

    c_owner = col("Owner")
    c_team  = col("Team Name")
    if not c_owner or not c_team:
        raise SystemExit("ERROR: docs/Team_Names.xlsx must have headers: Owner, Team Name")

    m: Dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        old = ws.cell(row=r, column=c_owner).value
        new = ws.cell(row=r, column=c_team).value
        old_s = "" if old is None else str(old).strip()
        new_s = "" if new is None else str(new).strip()
        if old_s and new_s:
            m[old_s] = new_s

    return m

def display_team(name: str, team_map: Dict[str, str]) -> str:
    s = (name or "").strip()
    if not s:
        return s
    if s == "Undrafted":
        return s
    return team_map.get(s, s)

# ----------------------------
# Load rosters.xlsx (bio fields)
# ----------------------------
def load_rosters(team_map: Dict[str, str]) -> Dict[str, dict]:
    """
    rosters.xlsx headers (per screenshot):
      Name, Order, Cost, Owner, Team, Height, Weight, Class, Position
    We'll treat Owner as "Team Name" (but display via Team_Names.xlsx mapping).
    """
    if not os.path.exists(ROSTERS_XLSX):
        raise SystemExit(f"ERROR: Missing {ROSTERS_XLSX}")

    wb = load_workbook(ROSTERS_XLSX, data_only=True)
    ws = wb.active

    headers = [("" if c.value is None else str(c.value).strip()) for c in ws[1]]
    headers_l = [h.lower() for h in headers]

    def col(*cands):
        for c in cands:
            if c.lower() in headers_l:
                return headers_l.index(c.lower()) + 1
        return None

    c_name = col("name", "player")
    if not c_name:
        raise SystemExit("ERROR: rosters.xlsx must have a 'Name' column.")

    c_cost   = col("cost")
    c_owner  = col("owner", "team name")
    c_team   = col("team")
    c_height = col("height", "ht")
    c_weight = col("weight", "wt")
    c_class  = col("class")
    c_pos    = col("position", "pos")

    def sval(r: int, c: Optional[int]) -> str:
        if c is None:
            return ""
        v = ws.cell(row=r, column=c).value
        return "" if v is None else str(v).strip()

    out: Dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        name = sval(r, c_name)
        if not name:
            continue
        key = norm_name(name)

        owner_raw = sval(r, c_owner)
        owner_disp = display_team(owner_raw, team_map)

        out[key] = {
            "Name": name,
            "Cost": sval(r, c_cost),
            "Team Name": owner_disp,   # DISPLAY name
            "Team": sval(r, c_team),
            "Height": sval(r, c_height),
            "Weight": sval(r, c_weight),
            "Class": sval(r, c_class),
            "Position": sval(r, c_pos),
        }

    return out

# ----------------------------
# Load Final_Players_PD*.html (pooh per PD + stat totals)
# ----------------------------
def load_final_player_data(cap_pd: Optional[int], team_map: Dict[str, str]):
    """
    Returns:
      max_pd
      pooh_by_player_pd[player_norm][pd] = pooh (int)  (only if player appears)
      played_by_player[player_norm] = set(pd) where player appears in box score
      agg_stats[player_norm] totals across included PDs:
         games, min, pts, reb, ast, stl, blk, to
      owner_by_player[player_norm] = DISPLAY team name (mapped)
    """
    files = []
    for fn in os.listdir(DOCS_DIR):
        n = pd_num_from_filename(fn)
        if n is None:
            continue
        if cap_pd is not None and n > cap_pd:
            continue
        files.append((n, fn))
    files.sort(key=lambda x: x[0])

    if not files:
        raise SystemExit("ERROR: No docs/Final_Players_PD*.html found.")

    max_pd = files[-1][0]

    pooh_by_player_pd: Dict[str, Dict[int, int]] = defaultdict(dict)
    played_by_player: Dict[str, set] = defaultdict(set)

    agg = defaultdict(lambda: {"games": 0, "min": 0.0, "pts": 0, "reb": 0, "ast": 0, "stl": 0, "blk": 0, "to": 0})
    owner_by_player: Dict[str, str] = {}

    for pd, fn in files:
        path = os.path.join(DOCS_DIR, fn)
        headers_l, rows = html_read_table(path)
        if not headers_l or not rows:
            continue

        i_owner  = idx(headers_l, "owner")
        i_player = idx(headers_l, "player")
        i_pooh   = idx(headers_l, "pooh")
        i_pts    = idx(headers_l, "pts")
        i_reb    = idx(headers_l, "reb")
        i_ast    = idx(headers_l, "ast")
        i_stl    = idx(headers_l, "stl")
        i_blk    = idx(headers_l, "blk")
        i_to     = idx(headers_l, "to")
        i_min    = idx(headers_l, "min")

        if i_player is None or i_pooh is None:
            continue

        for r in rows:
            if i_player >= len(r):
                continue
            pname = r[i_player]
            key = norm_name(pname)
            if not key:
                continue

            # Mark as played for this PD (player appears in box score)
            played_by_player[key].add(pd)

            pooh = safe_int(r[i_pooh]) if i_pooh < len(r) else 0
            pooh_by_player_pd[key][pd] = pooh

            # Count as a game played for that PD.
            agg[key]["games"] += 1
            if i_min is not None and i_min < len(r):
                agg[key]["min"] += safe_float(r[i_min])
            if i_pts is not None and i_pts < len(r):
                agg[key]["pts"] += safe_int(r[i_pts])
            if i_reb is not None and i_reb < len(r):
                agg[key]["reb"] += safe_int(r[i_reb])
            if i_ast is not None and i_ast < len(r):
                agg[key]["ast"] += safe_int(r[i_ast])
            if i_stl is not None and i_stl < len(r):
                agg[key]["stl"] += safe_int(r[i_stl])
            if i_blk is not None and i_blk < len(r):
                agg[key]["blk"] += safe_int(r[i_blk])
            if i_to is not None and i_to < len(r):
                agg[key]["to"] += safe_int(r[i_to])

            # Owner/team name (map for display)
            if i_owner is not None and i_owner < len(r):
                ow_raw = r[i_owner].strip()
                if ow_raw:
                    owner_by_player[key] = display_team(ow_raw, team_map)

    return max_pd, pooh_by_player_pd, played_by_player, agg, owner_by_player

# ----------------------------
# Write HTML
# ----------------------------
NUM_COLS = {"Cost", "Min/G", "Avg", "Total", "PPG", "R/G", "A/G", "B/G", "S/G", "T/G"}

def write_html(out_path: str, cols: List[str], rows: List[Dict[str, str]], title: str):
    with open(out_path, "w", encoding="utf-8") as out:
        out.write("<!doctype html><html><head><meta charset='utf-8'>")
        out.write(f"<title>{title}</title>")
        out.write(
            "<style>"
            "body{font-family:Arial}"
            "table{border-collapse:collapse;font-size:14px}"
            "th,td{border:1px solid #ccc;padding:4px 6px}"
            "th{background:#eee}"
            "td.num{text-align:right}"
            "</style>"
        )
        out.write("</head><body>")
        out.write(f"<h2 style='text-align:center'>{title}</h2>")

        out.write("<table><thead><tr>")
        for c in cols:
            out.write(f"<th>{c}</th>")
        out.write("</tr></thead><tbody>")

        for r in rows:
            out.write("<tr>")
            for c in cols:
                v = r.get(c, "")
                is_num = (c in NUM_COLS) or c.isdigit()
                cls = " class='num'" if is_num else ""
                out.write(f"<td{cls}>{v}</td>")
            out.write("</tr>")

        out.write("</tbody></table></body></html>")

    print(f"Wrote: {out_path}")

# ----------------------------
# Main
# ----------------------------
def main():
    cap_pd = parse_cap_pd(sys.argv)

    team_map = load_team_name_map()
    if team_map:
        print(f"Loaded Team_Names mapping entries: {len(team_map)}")
    else:
        print("No Team_Names mapping loaded (using names as-is).")

    rosters = load_rosters(team_map)
    max_pd, pooh_by_player_pd, played_by_player, agg, owner_by_player = load_final_player_data(cap_pd, team_map)

    # Columns EXACTLY as you requested (with spelled-out Height/Weight)
    fixed_cols = ["Team Name", "Cost", "Name", "Team", "Height", "Weight", "Class", "Position", "Min/G", "Avg", "Total"]
    pd_cols = [str(i) for i in range(1, max_pd + 1)]
    tail_cols = ["PPG", "R/G", "A/G", "B/G", "S/G", "T/G"]
    cols = fixed_cols + pd_cols + tail_cols

    rows_out: List[Dict[str, str]] = []

    # Build rows from roster list (keeps every rostered player even if they never played)
    for key, info in rosters.items():
        played_set = played_by_player.get(key, set())

        # PD cells: blank if not played; otherwise number (including 0)
        total_pooh = 0
        played_count = 0
        for pd in range(1, max_pd + 1):
            if pd in played_set:
                v = pooh_by_player_pd.get(key, {}).get(pd, 0)
                total_pooh += int(v)
                played_count += 1

        avg_pooh = (total_pooh / played_count) if played_count > 0 else 0.0

        g = agg.get(key, {"games": 0, "min": 0.0, "pts": 0, "reb": 0, "ast": 0, "stl": 0, "blk": 0, "to": 0})
        games = g["games"] if g else 0

        def per_game(n: float) -> float:
            return (n / games) if games > 0 else 0.0

        min_g = per_game(g["min"])
        ppg   = per_game(g["pts"])
        rpg   = per_game(g["reb"])
        apg   = per_game(g["ast"])
        bpg   = per_game(g["blk"])
        spg   = per_game(g["stl"])
        tpg   = per_game(g["to"])

        # Prefer owner from Final files; else roster value. Both already display-mapped.
        team_name_disp = owner_by_player.get(key) or info.get("Team Name", "")

        row = {
            "Team Name": team_name_disp,
            "Cost": info.get("Cost", ""),
            "Name": info.get("Name", ""),
            "Team": info.get("Team", ""),
            "Height": info.get("Height", ""),
            "Weight": info.get("Weight", ""),
            "Class": info.get("Class", ""),
            "Position": info.get("Position", ""),
            "Min/G": f"{min_g:.1f}",
            "Avg": f"{avg_pooh:.2f}",
            "Total": str(total_pooh),
            "PPG": f"{ppg:.2f}",
            "R/G": f"{rpg:.2f}",
            "A/G": f"{apg:.2f}",
            "B/G": f"{bpg:.2f}",
            "S/G": f"{spg:.2f}",
            "T/G": f"{tpg:.2f}",
        }

        for pd in range(1, max_pd + 1):
            if pd in played_set:
                row[str(pd)] = str(pooh_by_player_pd.get(key, {}).get(pd, 0))
            else:
                row[str(pd)] = ""  # BLANK if not played / not in box score

        rows_out.append(row)

    # 1) Player_Pooh_Summary.html: sort by Avg desc (new rules), then Total desc, then Name
    rows_players = sorted(
        rows_out,
        key=lambda r: (-safe_float(r.get("Avg", "0")), -safe_int(r.get("Total", "0")), r.get("Name", ""))
    )
    write_html(OUT_PLAYER, cols, rows_players, title="Player Pooh Summary")

    # 2) Pooh_Summary_By_Team.html: sort by Team Name (asc), then Avg desc, then Name
    def sort_key_by_team_then_avg(r):
        team = (r.get("Team Name") or "").strip().lower()
        avg = safe_float(r.get("Avg", 0))
        name = (r.get("Name") or "")
        return (team, -avg, name)

    rows_by_team = sorted(rows_out, key=sort_key_by_team_then_avg)
    write_html(OUT_BY_TEAM, cols, rows_by_team, title="Pooh Summary By Team")

if __name__ == "__main__":
    main()
