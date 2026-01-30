import os
import re
import sys
import html
from collections import defaultdict
from typing import Dict, List, Optional, Tuple, Set

from bs4 import BeautifulSoup
from openpyxl import load_workbook

REPO_ROOT = os.path.join(os.path.dirname(__file__), "..")
APP_DIR   = os.path.join(REPO_ROOT, "app")
DOCS_DIR  = os.path.join(REPO_ROOT, "docs")

ROSTERS_XLSX    = os.path.join(APP_DIR, "rosters.xlsx")
TEAM_NAMES_XLSX = os.path.join(DOCS_DIR, "Team_Names.xlsx")

OUT_DIR = os.path.join(DOCS_DIR, "Teams")


# ----------------------------
# Helpers
# ----------------------------
def parse_cap_pd(argv) -> Optional[int]:
    # optional arg: PD7
    if len(argv) < 2:
        return None
    s = argv[1].strip().upper()
    m = re.fullmatch(r"PD(\d+)", s)
    if not m:
        raise SystemExit("Usage: python app/build_team_pages.py [PD7]")
    return int(m.group(1))


def norm_name(name: str) -> str:
    s = (name or "").lower()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\b(jr|sr|ii|iii|iv)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def safe_int(x) -> int:
    try:
        return int(str(x).strip())
    except:
        return 0


def safe_float(x) -> float:
    try:
        return float(str(x).strip())
    except:
        return 0.0


def sanitize_team_filename(team_name: str) -> str:
    s = (team_name or "").strip()
    s = s.replace(" ", "_")
    s = re.sub(r"[^A-Za-z0-9_\-]", "", s)
    if not s:
        s = "Team"
    return s + ".html"


def pd_num_from_players_filename(fn: str) -> Optional[int]:
    m = re.search(r"Final_Players_PD(\d+)\.html$", fn)
    return int(m.group(1)) if m else None


# ----------------------------
# Team name mapping (Owner -> Team Name)
# ----------------------------
def load_team_name_map() -> Dict[str, str]:
    if not os.path.exists(TEAM_NAMES_XLSX):
        raise SystemExit(f"ERROR: Missing {TEAM_NAMES_XLSX}")

    wb = load_workbook(TEAM_NAMES_XLSX, data_only=True)
    ws = wb.active

    headers = [("" if c.value is None else str(c.value).strip()) for c in ws[1]]
    headers_l = [h.lower() for h in headers]

    def col(*cands):
        for c in cands:
            if c.lower() in headers_l:
                return headers_l.index(c.lower()) + 1
        return None

    c_owner = col("owner")
    c_team  = col("team name", "team")
    if not c_owner or not c_team:
        raise SystemExit("ERROR: Team_Names.xlsx must have headers: Owner, Team Name")

    out = {}
    for r in range(2, ws.max_row + 1):
        owner = ws.cell(row=r, column=c_owner).value
        team  = ws.cell(row=r, column=c_team).value
        owner = "" if owner is None else str(owner).strip()
        team  = "" if team is None else str(team).strip()
        if owner and team:
            out[owner] = team

    return out


def display_team_name(owner_old: str, team_map: Dict[str, str]) -> str:
    return team_map.get(owner_old, owner_old)


# ----------------------------
# Rosters loader (needs Draft Order + bio fields)
# ----------------------------
def load_rosters() -> Dict[str, dict]:
    if not os.path.exists(ROSTERS_XLSX):
        raise SystemExit(f"ERROR: Missing {ROSTERS_XLSX}")

    wb = load_workbook(ROSTERS_XLSX, data_only=True)
    ws = wb.active

    headers = [("" if c.value is None else str(c.value).strip()) for c in ws[1]]
    headers_l = [h.lower() for h in headers]

    def col(*cands):
        for c in cands:
            if c.lower() in headers_l:
                return headers_l.index(c.lower()) + 1
        return None

    c_name  = col("name", "player")
    c_order = col("order", "draft order")
    c_cost  = col("cost")
    c_owner = col("owner", "team name")
    c_team  = col("team")
    c_pos   = col("position", "pos")
    c_ht    = col("height", "ht")
    c_wt    = col("weight", "wt")
    c_class = col("class")

    if not c_name or not c_owner:
        raise SystemExit("ERROR: rosters.xlsx must have at least Name and Owner columns.")

    def sval(r: int, c: Optional[int]) -> str:
        if c is None:
            return ""
        v = ws.cell(row=r, column=c).value
        return "" if v is None else str(v).strip()

    out: Dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        name = sval(r, c_name)
        if not name:
            continue
        key = norm_name(name)

        out[key] = {
            "Name": name,
            "Draft Order": sval(r, c_order),
            "Cost": sval(r, c_cost),
            "Owner": sval(r, c_owner),      # old owner/team label
            "Team": sval(r, c_team),
            "Position": sval(r, c_pos),
            "Height": sval(r, c_ht),
            "Weight": sval(r, c_wt),
            "Class": sval(r, c_class),
        }

    return out


# ----------------------------
# Parse Final_Players_PD*.html and detect "starter" via:
#  - your older files: td class='start' OR started_today == Yes
#  - your newer files: bold text (<b>/<strong>) in the row cells
# ----------------------------
def parse_final_players_pd_file(path: str) -> Tuple[List[str], List[dict]]:
    """
    Returns:
      headers_l
      rows: list of dicts with keys including:
        owner, started_today, player, pooh, pts, reb, ast, stl, blk, to, min
      plus: __is_starter (bool) detected via:
        - started_today == Yes
        - td class 'start'
        - any <b> or <strong> within the row's td cells
    """
    with open(path, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f.read(), "html.parser")

    table = soup.find("table")
    if not table:
        return [], []

    # headers
    ths = table.find_all("th")
    headers = [th.get_text(strip=True) for th in ths]
    headers_l = [h.lower() for h in headers]

    rows_out: List[dict] = []
    trs = table.find_all("tr")
    for tr in trs[1:]:
        tds = tr.find_all("td")
        if not tds:
            continue

        # starter signals
        is_starter_by_css = any(("start" in (td.get("class") or [])) for td in tds)
        is_starter_by_bold = any((td.find("b") is not None) or (td.find("strong") is not None) for td in tds)

        values = [td.get_text(strip=True) for td in tds]
        row = {}
        for i, h in enumerate(headers_l):
            if i < len(values):
                row[h] = values[i]

        started_today_txt = (row.get("started_today") or row.get("started today") or "").strip().lower()
        is_starter_by_col = started_today_txt in ("yes", "y", "true", "1")

        row["__is_starter"] = bool(is_starter_by_css or is_starter_by_col or is_starter_by_bold)
        rows_out.append(row)

    return headers_l, rows_out


def load_final_player_data_and_actuals(cap_pd: Optional[int], team_map_rev: Dict[str, str]):
    """
    Returns:
      max_pd
      pooh_by_player_pd[player_norm][pd] = pooh
      agg_stats[player_norm] = totals across included PDs:
         games, min, pts, reb, ast, stl, blk, to
      actual_by_owner_pd[owner_old][pd] = sum(starter pooh for that PD
      starters_by_owner_pd[owner_old][pd] = set(player_norm) who were starters for that PD
    """
    files = []
    for fn in os.listdir(DOCS_DIR):
        n = pd_num_from_players_filename(fn)
        if n is None:
            continue
        if cap_pd is not None and n > cap_pd:
            continue
        files.append((n, fn))
    files.sort(key=lambda x: x[0])

    if not files:
        raise SystemExit("ERROR: No docs/Final_Players_PD*.html found.")

    max_pd = files[-1][0]

    pooh_by_player_pd: Dict[str, Dict[int, int]] = defaultdict(dict)
    agg = defaultdict(lambda: {"games": 0, "min": 0.0, "pts": 0, "reb": 0, "ast": 0, "stl": 0, "blk": 0, "to": 0})
    actual_by_owner_pd: Dict[str, Dict[int, int]] = defaultdict(dict)
    starters_by_owner_pd: Dict[str, Dict[int, Set[str]]] = defaultdict(lambda: defaultdict(set))

    for pd, fn in files:
        path = os.path.join(DOCS_DIR, fn)
        headers_l, rows = parse_final_players_pd_file(path)
        if not headers_l or not rows:
            continue

        for r in rows:
            owner_raw = (r.get("owner") or "").strip()
            player = (r.get("player") or "").strip()
            if not owner_raw or not player:
                continue

            # Normalize "owner" back to OLD owner name (so PDs with new team names still map)
            owner = team_map_rev.get(owner_raw, owner_raw)

            key = norm_name(player)
            if not key:
                continue

            pooh = safe_int(r.get("pooh", 0))
            pooh_by_player_pd[key][pd] = pooh

            # aggregate totals
            agg[key]["games"] += 1
            agg[key]["min"] += safe_float(r.get("min", 0.0))
            agg[key]["pts"] += safe_int(r.get("pts", 0))
            agg[key]["reb"] += safe_int(r.get("reb", 0))
            agg[key]["ast"] += safe_int(r.get("ast", 0))
            agg[key]["stl"] += safe_int(r.get("stl", 0))
            agg[key]["blk"] += safe_int(r.get("blk", 0))
            agg[key]["to"]  += safe_int(r.get("to", 0))

            # ACTUAL starters (bold / class / started_today)
            if r.get("__is_starter"):
                actual_by_owner_pd[owner][pd] = actual_by_owner_pd[owner].get(pd, 0) + pooh
                starters_by_owner_pd[owner][pd].add(key)

    return max_pd, pooh_by_player_pd, agg, actual_by_owner_pd, starters_by_owner_pd


# ----------------------------
# Lineup constraint for MAX: GGGFF or GGFFF
# ----------------------------
def classify_pos(pos: str) -> str:
    """
    Returns:
      "G"  = guard-only
      "F"  = forward/frontcourt-only (F or C)
      "GF" = flex (contains both G and F/C)
      ""   = unknown (treat as flex)
    """
    p = (pos or "").upper().replace(" ", "")
    if not p:
        return ""

    has_g = "G" in p
    has_f = ("F" in p) or ("C" in p)

    if has_g and has_f:
        return "GF"
    if has_g:
        return "G"
    if has_f:
        return "F"
    return ""


def best_valid_lineup_sum(player_items: List[Tuple[int, str]]) -> int:
    """
    FAST DP version.
    player_items = [(pooh, pos_class), ...]
      pos_class in {"G","F","GF",""}   ("" treated as GF)

    Need best 5-player sum where guards in {2,3}.
    """
    n = len(player_items)
    if n < 5:
        return 0

    NEG = -10**18

    # dp[picked][guards] = best sum achievable
    dp = [[NEG] * 4 for _ in range(6)]
    dp[0][0] = 0

    for pooh, pos_class in player_items:
        pc = (pos_class or "")
        if pc not in ("G", "F", "GF", ""):
            pc = ""

        ndp = [row[:] for row in dp]

        for picked in range(0, 5):
            for guards in range(0, 4):
                cur = dp[picked][guards]
                if cur == NEG:
                    continue

                # take as Guard
                if pc in ("G", "GF", ""):
                    np = picked + 1
                    ng = guards + 1
                    if ng <= 3:
                        nf = np - ng
                        if nf <= 3:
                            ndp[np][ng] = max(ndp[np][ng], cur + pooh)

                # take as Forward
                if pc in ("F", "GF", ""):
                    np = picked + 1
                    ng = guards
                    nf = np - ng
                    if nf <= 3:
                        ndp[np][ng] = max(ndp[np][ng], cur + pooh)

        dp = ndp

    best = max(dp[5][2], dp[5][3])
    return 0 if best == NEG else int(best)


# ----------------------------
# Per-game helpers
# ----------------------------
def per_game(n: float, games: int) -> float:
    return (n / games) if games > 0 else 0.0


# ----------------------------
# HTML output (team page)
# ----------------------------
def write_team_page(
    out_path: str,
    team_title: str,
    cols: List[str],
    rows: List[Dict[str, str]],
    pd_highlight_cells: Dict[Tuple[int, str], bool],  # (pd, player_key) -> highlight?
    actual_pd: Dict[int, int],
    max_pd_sums: Dict[int, int],
    total_actual: int,
    total_max: int,
):
    def esc(x) -> str:
        return html.escape("" if x is None else str(x))

    # Column widths tuned to fit “one page wide” better
    def col_width_px(c: str) -> int:
        if c.isdigit():
            return 28  # PD columns tight

        widths = {
            "Team Name": 70,
            "Name": 140,
            "Draft Order": 55,
            "Cost": 45,
            "Team": 45,
            "Position": 45,
            "Height": 45,
            "Weight": 55,
            "Class": 40,
            "Min/G": 45,
            "Total": 55,
            "Avg": 45,
            "PPG": 45,
            "R/G": 45,
            "A/G": 45,
            "B/G": 45,
            "S/G": 45,
            "T/G": 45,
        }
        return widths.get(c, 45)

    with open(out_path, "w", encoding="utf-8") as out:
        out.write("<!doctype html><html><head><meta charset='utf-8'>")
        out.write(f"<title>{esc(team_title)}</title>")
        out.write("""
<style>
body{font-family:Arial;background:#ffffff;margin:0;padding:0}
.wrapper{width:1100px;margin:12px auto;border:3px solid #000;background:#FFFFCC;padding:8px}
h1{font-size:20px;text-align:center;margin:0 0 8px 0}

/* fit on one page wide */
table{width:100%;border-collapse:collapse;font-size:11px;background:#ffffff;table-layout:fixed}
th,td{border:1px solid #333;padding:2px 3px;text-align:center;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
th{background:#C0C0C0;font-size:11px}

td.num{text-align:right}
td.blank{background:#000;color:#000}
td.hit{background:#00FFFF}

tr.totalrow td{background:#000;color:#000}
tr.totalrow td.keep{background:#ffffff;color:#000;font-weight:bold}
</style>
        """)

        out.write("</head><body><div class='wrapper'>")
        out.write(f"<h1>{esc(team_title)}</h1>")

        out.write("<table><colgroup>")
        for c in cols:
            out.write(f"<col style='width:{col_width_px(c)}px'>")
        out.write("</colgroup><thead><tr>")
        for c in cols:
            out.write(f"<th title='{esc(c)}'>{esc(c)}</th>")
        out.write("</tr></thead><tbody>")

        # data rows
        for r in rows:
            out.write("<tr>")
            player_key = r.get("__key", "")
            for c in cols:
                v = r.get(c, "")

                if c.isdigit():
                    pd = int(c)
                    if v == "":
                        out.write("<td class='blank'>&nbsp;</td>")
                    else:
                        cls = "hit" if pd_highlight_cells.get((pd, player_key), False) else ""
                        out.write(f"<td class='{cls}'>{esc(v)}</td>" if cls else f"<td>{esc(v)}</td>")
                    continue

                if c in {"Draft Order","Cost","Min/G","Total","Avg","PPG","R/G","A/G","B/G","S/G","T/G"}:
                    out.write(f"<td class='num'>{esc(v)}</td>")
                else:
                    out.write(f"<td title='{esc(v)}'>{esc(v)}</td>")
            out.write("</tr>")

        # Bottom: Actual row
        out.write("<tr class='totalrow'>")
        for c in cols:
            if c == "Total":
                out.write(f"<td class='keep num'>{total_actual}</td>")
            elif c == "Avg":
                out.write("<td class='keep'>Actual</td>")
            elif c.isdigit():
                pd = int(c)
                out.write(f"<td class='keep num'>{actual_pd.get(pd, 0)}</td>")
            else:
                out.write("<td>&nbsp;</td>")
        out.write("</tr>")

        # Bottom: Max row
        out.write("<tr class='totalrow'>")
        for c in cols:
            if c == "Total":
                out.write(f"<td class='keep num'>{total_max}</td>")
            elif c == "Avg":
                out.write("<td class='keep'>Max</td>")
            elif c.isdigit():
                pd = int(c)
                out.write(f"<td class='keep num'>{max_pd_sums.get(pd, 0)}</td>")
            else:
                out.write("<td>&nbsp;</td>")
        out.write("</tr>")

        out.write("</tbody></table></div></body></html>")

    print(f"Wrote: {out_path}")


# ----------------------------
# Main
# ----------------------------
def main():
    cap_pd = parse_cap_pd(sys.argv)

    team_map = load_team_name_map()
    team_map_rev = {v: k for k, v in team_map.items()}  # Team Name -> Owner (old)
    rosters = load_rosters()

    # Load player PD values + aggregates + ACTUAL starter totals + which starters they were
    max_pd, pooh_by_player_pd, agg, actual_by_owner_pd, starters_by_owner_pd = load_final_player_data_and_actuals(
        cap_pd, team_map_rev
    )

    os.makedirs(OUT_DIR, exist_ok=True)

    # Group rostered players by OLD owner name
    players_by_owner: Dict[str, List[str]] = defaultdict(list)
    for k, info in rosters.items():
        owner_old = info.get("Owner", "")
        players_by_owner[owner_old].append(k)

    # Build pages per owner
    for owner_old, player_keys in players_by_owner.items():
        team_name = display_team_name(owner_old, team_map)
        team_title = f"Sorted Summary Results For {team_name}"

        fixed_cols = [
            "Team Name","Name","Draft Order","Cost","Team","Position","Height","Weight","Class","Min/G","Total","Avg"
        ]
        pd_cols = [str(i) for i in range(1, max_pd + 1)]
        tail_cols = ["PPG","R/G","A/G","B/G","S/G","T/G"]
        cols = fixed_cols + pd_cols + tail_cols

        # Highlight = ACTUAL starters per PD (from Final_Players bold rows)
        pd_highlight_cells: Dict[Tuple[int, str], bool] = {}
        for pd in range(1, max_pd + 1):
            starter_keys = starters_by_owner_pd.get(owner_old, {}).get(pd, set())
            for k in starter_keys:
                pd_highlight_cells[(pd, k)] = True

        # Build player rows
        rows_out: List[Dict[str, str]] = []
        for k in player_keys:
            info = rosters.get(k, {})
            g = agg.get(k, {"games": 0, "min": 0.0, "pts": 0, "reb": 0, "ast": 0, "stl": 0, "blk": 0, "to": 0})
            games = int(g.get("games", 0))

            # PD values: BLANK if player not in boxscore at all for that PD
            played_pd_count = 0
            total_pooh = 0
            pd_vals: List[Optional[int]] = []
            for pd in range(1, max_pd + 1):
                if pd in pooh_by_player_pd.get(k, {}):
                    v = int(pooh_by_player_pd[k][pd])
                    pd_vals.append(v)
                    total_pooh += v
                    played_pd_count += 1
                else:
                    pd_vals.append(None)

            avg_pooh = (total_pooh / played_pd_count) if played_pd_count > 0 else 0.0

            min_g = per_game(float(g.get("min", 0.0)), games)
            ppg   = per_game(float(g.get("pts", 0)), games)
            rpg   = per_game(float(g.get("reb", 0)), games)
            apg   = per_game(float(g.get("ast", 0)), games)
            bpg   = per_game(float(g.get("blk", 0)), games)
            spg   = per_game(float(g.get("stl", 0)), games)
            tpg   = per_game(float(g.get("to", 0)), games)

            row = {
                "__key": k,
                "Team Name": team_name,
                "Name": info.get("Name", ""),
                "Draft Order": info.get("Draft Order", ""),
                "Cost": info.get("Cost", ""),
                "Team": info.get("Team", ""),
                "Position": info.get("Position", ""),
                "Height": info.get("Height", ""),
                "Weight": info.get("Weight", ""),
                "Class": info.get("Class", ""),
                "Min/G": f"{min_g:.1f}" if games > 0 else "",
                "Total": str(total_pooh) if played_pd_count > 0 else "0",
                "Avg": f"{avg_pooh:.2f}" if played_pd_count > 0 else "0.00",
                "PPG": f"{ppg:.1f}" if games > 0 else "",
                "R/G": f"{rpg:.1f}" if games > 0 else "",
                "A/G": f"{apg:.1f}" if games > 0 else "",
                "B/G": f"{bpg:.1f}" if games > 0 else "",
                "S/G": f"{spg:.1f}" if games > 0 else "",
                "T/G": f"{tpg:.1f}" if games > 0 else "",
            }

            for pd in range(1, max_pd + 1):
                v = pd_vals[pd - 1]
                row[str(pd)] = "" if v is None else str(v)

            rows_out.append(row)

        # Sort like your view: Avg desc, Total desc, then Name
        rows_out.sort(key=lambda r: (-safe_float(r.get("Avg", 0)), -safe_int(r.get("Total", 0)), r.get("Name", "")))

        # ACTUAL per PD: from Final_Players starters (bold)
        actual_pd: Dict[int, int] = {}
        for pd in range(1, max_pd + 1):
            actual_pd[pd] = int(actual_by_owner_pd.get(owner_old, {}).get(pd, 0))
        total_actual = sum(actual_pd.values())

        # MAX per PD: best valid 5-man lineup under (2-3 G) and (2-3 F/C)
        max_pd_sums: Dict[int, int] = {}
        for pd in range(1, max_pd + 1):
            items: List[Tuple[int, str]] = []
            for k in player_keys:
                if pd not in pooh_by_player_pd.get(k, {}):
                    continue
                pooh = int(pooh_by_player_pd[k][pd])
                pos_class = classify_pos(rosters.get(k, {}).get("Position", ""))
                items.append((pooh, pos_class))
            max_pd_sums[pd] = best_valid_lineup_sum(items)

        total_max = sum(max_pd_sums.values())

        out_path = os.path.join(OUT_DIR, sanitize_team_filename(team_name))
        write_team_page(
            out_path=out_path,
            team_title=team_title,
            cols=cols,
            rows=rows_out,
            pd_highlight_cells=pd_highlight_cells,
            actual_pd=actual_pd,
            max_pd_sums=max_pd_sums,
            total_actual=total_actual,  # Total column in "Actual" row
            total_max=total_max,        # Total column in "Max" row
        )


if __name__ == "__main__":
    main()
