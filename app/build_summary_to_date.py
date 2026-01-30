import os
import re
import sys
from collections import defaultdict
from bs4 import BeautifulSoup
from openpyxl import load_workbook

DOCS_DIR = os.path.join(os.path.dirname(__file__), "..", "docs")
TEAM_NAMES_XLSX = os.path.join(DOCS_DIR, "Team_Names.xlsx")


def parse_cap_pd(argv) -> int | None:
    # optional: PD7
    if len(argv) < 2:
        return None
    s = argv[1].strip().upper()
    m = re.fullmatch(r"PD(\d+)", s)
    if not m:
        raise SystemExit("Usage: python app/build_summary_to_date.py [PD7]")
    return int(m.group(1))


def pd_num_from_filename(fn: str) -> int | None:
    m = re.search(r"Final_Owners_PD(\d+)\.html$", fn)
    return int(m.group(1)) if m else None


def read_owner_totals_from_final_owners_html(path: str) -> dict[str, int]:
    """
    Reads docs/Final_Owners_PDx.html table like:
      Owner | Starter Pooh Total | Starters Count So Far
    Returns {owner_cell_text: starter_pooh_total}.
    """
    with open(path, "r", encoding="utf-8") as f:
        soup = BeautifulSoup(f.read(), "html.parser")

    table = soup.find("table")
    if not table:
        return {}

    rows = table.find_all("tr")
    out: dict[str, int] = {}

    for tr in rows[1:]:
        tds = tr.find_all("td")
        if len(tds) < 2:
            continue
        owner = tds[0].get_text(strip=True)
        total_txt = tds[1].get_text(strip=True)
        try:
            out[owner] = int(str(total_txt).strip())
        except:
            out[owner] = 0

    return out


# ----------------------------
# Team Names mapping
# ----------------------------
def load_team_name_map() -> dict[str, str]:
    """
    docs/Team_Names.xlsx
      Column A: Owner (old name)
      Column B: Team Name (new name)
    Headers: Owner, Team Name
    """
    if not os.path.exists(TEAM_NAMES_XLSX):
        # If the file isn't present yet, just behave like "no mapping"
        return {}

    wb = load_workbook(TEAM_NAMES_XLSX, data_only=True)
    ws = wb.active

    headers = [("" if c.value is None else str(c.value).strip()) for c in ws[1]]
    headers_l = [h.lower() for h in headers]

    def col(*cands):
        for c in cands:
            c = c.lower()
            if c in headers_l:
                return headers_l.index(c) + 1
        return None

    c_owner = col("owner")
    c_team  = col("team name", "team")

    if not c_owner or not c_team:
        raise SystemExit("ERROR: Team_Names.xlsx must have headers: Owner, Team Name")

    out: dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        ow = ws.cell(row=r, column=c_owner).value
        tn = ws.cell(row=r, column=c_team).value
        owner = "" if ow is None else str(ow).strip()
        team  = "" if tn is None else str(tn).strip()
        if owner and team:
            out[owner] = team

    return out


def canon_owner_key(s: str) -> str:
    # normalize keys so "G-Flop" and "g flop" don't split
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("–", "-").replace("—", "-")
    return s


def build_owner_to_team_normalizer(team_map: dict[str, str]):
    """
    Returns normalize_owner_to_team(owner_raw) -> canonical team display name.
    Handles:
      - old Owner -> Team Name (via Team_Names.xlsx)
      - already-a-Team-Name -> itself (so it still groups correctly)
    """
    team_values = set(team_map.values())

    def normalize_owner_to_team(owner_raw: str) -> str:
        s = (owner_raw or "").strip()
        if not s:
            return s
        if s == "Undrafted":
            return s

        # If it's an old owner name, map to Team Name
        if s in team_map:
            return team_map[s]

        # If it's already one of the Team Name values, keep it
        if s in team_values:
            return s

        # Otherwise leave as-is
        return s

    return normalize_owner_to_team


def main():
    cap_pd = parse_cap_pd(sys.argv)

    # Load owner->team mapping and build normalizer
    team_map = load_team_name_map()
    normalize_owner_to_team = build_owner_to_team_normalizer(team_map)

    # Find Final_Owners_PD*.html
    pd_files: list[tuple[int, str]] = []
    for fn in os.listdir(DOCS_DIR):
        n = pd_num_from_filename(fn)
        if n is None:
            continue
        if cap_pd is not None and n > cap_pd:
            continue
        pd_files.append((n, fn))

    pd_files.sort(key=lambda x: x[0])  # PD1..PDN

    if not pd_files:
        raise SystemExit("No Final_Owners_PD*.html files found in docs/")

    max_pd = pd_files[-1][0]

    # per_team_per_pd[key][pd] = points for that PD
    per_team_per_pd: dict[str, dict[int, int]] = defaultdict(dict)
    display_name_by_key: dict[str, str] = {}
    keys_set = set()

    for pd, fn in pd_files:
        path = os.path.join(DOCS_DIR, fn)
        totals = read_owner_totals_from_final_owners_html(path)

        for owner_raw, v in totals.items():
            # IMPORTANT: normalize whatever is in the cell to the canonical TEAM NAME first
            team_name = normalize_owner_to_team(owner_raw)
            k = canon_owner_key(team_name)

            keys_set.add(k)
            per_team_per_pd[k][pd] = int(v)

            if k not in display_name_by_key:
                display_name_by_key[k] = team_name

    keys = sorted(list(keys_set), key=lambda k: display_name_by_key.get(k, k))

    # Totals + avg
    pd_list = [pd for pd, _ in pd_files]
    completed_pd_count = len(pd_list)

    team_total: dict[str, int] = {}
    team_avg: dict[str, float] = {}

    for k in keys:
        pd_scores = [per_team_per_pd[k].get(pd, 0) for pd in pd_list]
        total = sum(pd_scores)
        team_total[k] = total
        team_avg[k] = (total / completed_pd_count) if completed_pd_count > 0 else 0.0

    # Sort by Total Pooh descending, then Team Name
    keys_sorted = sorted(keys, key=lambda k: (-team_total.get(k, 0), display_name_by_key.get(k, k)))

    # Reference totals for Out Of 1st/2nd/3rd
    top1 = team_total.get(keys_sorted[0], 0) if len(keys_sorted) >= 1 else 0
    top2 = team_total.get(keys_sorted[1], top1) if len(keys_sorted) >= 2 else top1
    top3 = team_total.get(keys_sorted[2], top2) if len(keys_sorted) >= 3 else top2

    # Write SummaryToDate.html
    out_path = os.path.join(DOCS_DIR, "SummaryToDate.html")

    with open(out_path, "w", encoding="utf-8") as out:
        out.write("<!doctype html><html><head><meta charset='utf-8'>")
        out.write("<title>Sorted League Results</title>")
        out.write(
            "<style>"
            "body{font-family:Arial}"
            "table{border-collapse:collapse;font-size:14px}"
            "th,td{border:1px solid #ccc;padding:4px 6px}"
            "th{background:#eee}"
            "td.num{text-align:right}"
            "</style>"
        )
        out.write("</head><body>")
        out.write("<h2 style='text-align:center'>Sorted League Results</h2>")

        out.write("<table><thead><tr>")
        out.write("<th>Team Name</th>")
        out.write("<th>Total Pooh</th>")
        out.write("<th>Out Of 1st</th>")
        out.write("<th>Out Of 2nd</th>")
        out.write("<th>Out Of 3rd</th>")

        for pd in range(1, max_pd + 1):
            out.write(f"<th>{pd}</th>")

        out.write("<th>Avg Pooh Per Completed PD</th>")
        out.write("<th>Sum of Avgs, Top 5 Eligible</th>")  # blank for now
        out.write("</tr></thead><tbody>")

        for k in keys_sorted:
            total = team_total.get(k, 0)
            out1 = max(0, top1 - total)
            out2 = max(0, top2 - total)
            out3 = max(0, top3 - total)

            team_display = display_name_by_key.get(k, "")

            out.write("<tr>")
            out.write(f"<td>{team_display}</td>")
            out.write(f"<td class='num'>{total}</td>")
            out.write(f"<td class='num'>{out1}</td>")
            out.write(f"<td class='num'>{out2}</td>")
            out.write(f"<td class='num'>{out3}</td>")

            for pd in range(1, max_pd + 1):
                out.write(f"<td class='num'>{per_team_per_pd[k].get(pd, 0)}</td>")

            out.write(f"<td class='num'>{team_avg.get(k, 0.0):.2f}</td>")

            # Blank column on purpose
            out.write("<td class='num'></td>")

            out.write("</tr>")

        out.write("</tbody></table></body></html>")

    print(f"Wrote: {out_path}")


if __name__ == "__main__":
    main()
