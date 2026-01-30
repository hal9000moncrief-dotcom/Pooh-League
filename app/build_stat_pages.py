import os
import re
import sys
import time
import random
import html
import requests
from datetime import datetime, timedelta, date as dt_date
from typing import Dict, List, Optional, Tuple
from collections import defaultdict

from openpyxl import load_workbook
from zoneinfo import ZoneInfo

# ----------------------------
# Paths / Config
# ----------------------------
REPO_ROOT = os.path.join(os.path.dirname(__file__), "..")
APP_DIR   = os.path.join(REPO_ROOT, "app")
DOCS_DIR  = os.path.join(REPO_ROOT, "docs")

PD_XLSX      = os.path.join(APP_DIR, "PD.xlsx")
ROSTERS_XLSX = os.path.join(APP_DIR, "rosters.xlsx")

TEAM_NAMES_XLSX = os.path.join(DOCS_DIR, "Team_Names.xlsx")

LOCAL_TZ = ZoneInfo("America/Chicago")
UTC_TZ   = ZoneInfo("UTC")

BASE = "https://site.api.espn.com/apis/site/v2/sports/basketball/mens-college-basketball"

HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept": "application/json,text/plain,*/*",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.espn.com/",
    "Connection": "keep-alive",
}

SESSION = requests.Session()
SESSION.headers.update(HEADERS)

BASE_DELAY = 0.25
JITTER     = 0.25
MAX_RETRIES = 6
TIMEOUT     = 30


# ----------------------------
# Utility helpers
# ----------------------------
def polite_sleep():
    time.sleep(BASE_DELAY + random.random() * JITTER)

def get_json(url: str) -> dict:
    last_err = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = SESSION.get(url, timeout=TIMEOUT)
            r.raise_for_status()
            polite_sleep()
            return r.json()
        except Exception as e:
            last_err = e
            time.sleep((0.7 ** attempt) + random.random() * 0.7)
    raise RuntimeError(f"Failed after retries: {url}\nLast error: {last_err}")

def safe_int(v) -> int:
    try:
        return int(str(v).strip())
    except:
        return 0

def safe_float(v) -> float:
    try:
        return float(str(v).strip())
    except:
        return 0.0

def parse_made_attempt(s: str) -> Tuple[int, int]:
    try:
        a, b = str(s).split("-")
        return int(a), int(b)
    except:
        return 0, 0

def to_minutes(v) -> float:
    if v is None:
        return 0.0
    s = str(v).strip()
    if not s or s == "--":
        return 0.0
    if ":" in s:
        try:
            mm, ss = s.split(":")
            return int(mm) + int(ss) / 60.0
        except:
            return 0.0
    return safe_float(s)

def norm_name(name: str) -> str:
    s = (name or "").lower()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\b(jr|sr|ii|iii|iv)\b", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def parse_yyyymmdd(s: str) -> datetime:
    s = (s or "").strip()
    if not re.fullmatch(r"\d{8}", s):
        raise ValueError("Date must be YYYYMMDD (8 digits).")
    dt = datetime.strptime(s, "%Y%m%d")
    return dt.replace(tzinfo=LOCAL_TZ)

def fmt_yyyymmdd(dt: datetime) -> str:
    return dt.strftime("%Y%m%d")

def event_local_yyyymmdd(e: dict) -> str:
    iso = (e.get("date") or "").strip()
    if not iso:
        return ""
    if iso.endswith("Z"):
        iso = iso[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(iso)
    except Exception:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=UTC_TZ)
    return dt.astimezone(LOCAL_TZ).strftime("%Y%m%d")


# ----------------------------
# Team name map (Owner -> Team Name)
# ----------------------------
def load_team_name_map() -> Dict[str, str]:
    """
    docs/Team_Names.xlsx
      Column headers: Owner, Team Name
      Owner = old name
      Team Name = new display name
    """
    if not os.path.exists(TEAM_NAMES_XLSX):
        print(f"NOTE: Missing {TEAM_NAMES_XLSX}. Team Name will display as-is.")
        return {}

    wb = load_workbook(TEAM_NAMES_XLSX, data_only=True)
    ws = wb.active

    headers = [("" if c.value is None else str(c.value).strip()) for c in ws[1]]
    headers_l = [h.lower() for h in headers]

    def col(name: str) -> Optional[int]:
        n = name.lower()
        if n in headers_l:
            return headers_l.index(n) + 1
        return None

    c_owner = col("Owner")
    c_team  = col("Team Name")
    if not c_owner or not c_team:
        raise SystemExit("ERROR: docs/Team_Names.xlsx must have headers: Owner, Team Name")

    m: Dict[str, str] = {}
    for r in range(2, ws.max_row + 1):
        old = ws.cell(row=r, column=c_owner).value
        new = ws.cell(row=r, column=c_team).value
        old_s = "" if old is None else str(old).strip()
        new_s = "" if new is None else str(new).strip()
        if old_s and new_s:
            m[old_s] = new_s

    print(f"Loaded Team_Names mapping entries: {len(m)}")
    return m

def display_team_name(old_owner: str, team_map: Dict[str, str]) -> str:
    s = (old_owner or "").strip()
    if not s:
        return s
    if s == "Undrafted":
        return s
    return team_map.get(s, s)


# ----------------------------
# PD.xlsx loader (ROBUST)
# ----------------------------
def _cell_to_date_yyyymmdd(cell_value) -> Optional[str]:
    if cell_value is None:
        return None

    if isinstance(cell_value, datetime):
        return cell_value.strftime("%Y%m%d")
    if isinstance(cell_value, dt_date):
        return cell_value.strftime("%Y%m%d")

    s = str(cell_value).strip()
    if not s:
        return None

    s2 = re.sub(r"\s+", "", s)
    digits = re.sub(r"\D", "", s2)

    if len(digits) == 8:
        first4 = int(digits[0:4])
        if 1900 <= first4 <= 2100:
            yyyy, mm, dd = digits[0:4], digits[4:6], digits[6:8]
        else:
            mm, dd, yyyy = digits[0:2], digits[2:4], digits[4:8]
        try:
            dt = datetime(int(yyyy), int(mm), int(dd), tzinfo=LOCAL_TZ)
            return dt.strftime("%Y%m%d")
        except:
            return None

    for fmt in ("%m/%d/%Y", "%m-%d-%Y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            dt = datetime.strptime(s2, fmt).replace(tzinfo=LOCAL_TZ)
            return dt.strftime("%Y%m%d")
        except:
            pass

    return None

def _cell_to_pd_num(cell_value) -> Optional[int]:
    if cell_value is None:
        return None
    try:
        if isinstance(cell_value, int):
            n = int(cell_value)
            return n if 1 <= n <= 500 else None
        if isinstance(cell_value, float):
            n = int(cell_value)
            if abs(cell_value - n) < 1e-9 and 1 <= n <= 500:
                return n
            return None
    except:
        pass

    s = str(cell_value).strip()
    if not s:
        return None
    s = re.sub(r"\.0$", "", s)
    if re.fullmatch(r"\d{1,3}", s):
        n = int(s)
        return n if 1 <= n <= 500 else None
    return None

def load_pd_map(pd_xlsx: str) -> Dict[int, str]:
    if not os.path.exists(pd_xlsx):
        raise SystemExit(f"ERROR: Missing {pd_xlsx}")

    wb = load_workbook(pd_xlsx, data_only=True)
    ws = wb.active

    out: Dict[int, str] = {}
    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 10) + 1)]

        pd_num = None
        d_yyyymmdd = None

        for v in row_vals:
            n = _cell_to_pd_num(v)
            if n is not None:
                pd_num = n
                break

        for v in row_vals:
            d = _cell_to_date_yyyymmdd(v)
            if d is not None:
                d_yyyymmdd = d
                break

        if pd_num is None or d_yyyymmdd is None:
            continue

        out[pd_num] = d_yyyymmdd

    if not out:
        sample = []
        for r in range(1, min(ws.max_row, 12) + 1):
            sample.append([ws.cell(row=r, column=c).value for c in range(1, min(ws.max_column, 6) + 1)])
        print("DEBUG PD.xlsx first rows (first 6 cols):")
        for i, row in enumerate(sample, start=1):
            print(f"Row {i}: {row}")
        raise SystemExit(f"ERROR: No PD rows found in {pd_xlsx}")

    return out


# ----------------------------
# Rosters loader (includes Cost/Height/Weight/Class/Pos)
# ----------------------------
def load_rosters(rosters_xlsx: str) -> Dict[str, dict]:
    if not os.path.exists(rosters_xlsx):
        raise SystemExit(f"ERROR: Missing {rosters_xlsx}")

    wb = load_workbook(rosters_xlsx, data_only=True)
    ws = wb.active

    headers = [("" if c.value is None else str(c.value).strip()) for c in ws[1]]
    headers_l = [h.lower() for h in headers]

    def col(*cands):
        for c in cands:
            if c.lower() in headers_l:
                return headers_l.index(c.lower()) + 1
        return None

    c_name   = col("name", "player")
    if not c_name:
        raise SystemExit("ERROR: rosters.xlsx must have a 'Name' column.")

    c_owner  = col("owner", "team name")
    c_team   = col("team")
    c_cost   = col("cost")
    c_height = col("height", "ht")
    c_weight = col("weight", "wt")
    c_class  = col("class")
    c_pos    = col("position", "pos")

    def sval(r: int, c: Optional[int]) -> str:
        if c is None:
            return ""
        v = ws.cell(row=r, column=c).value
        return "" if v is None else str(v).strip()

    out: Dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        name = sval(r, c_name)
        if not name:
            continue

        key = norm_name(name)
        out[key] = {
            "Name": name,
            "Team Name": sval(r, c_owner),  # old owner name
            "Team": sval(r, c_team),
            "Cost": sval(r, c_cost),
            "Height": sval(r, c_height),
            "Weight": sval(r, c_weight),
            "Class": sval(r, c_class),
            "Pos": sval(r, c_pos),
        }

    return out


# ----------------------------
# ESPN fetch: SEC events for a given local date (yyyymmdd)
# ----------------------------
def get_sec_events(date_yyyymmdd: str) -> List[dict]:
    url = f"{BASE}/scoreboard?dates={date_yyyymmdd}&groups=23&limit=500"
    data = get_json(url)
    events = data.get("events", []) or []
    return [e for e in events if event_local_yyyymmdd(e) == date_yyyymmdd]


# ----------------------------
# Boxscore parsing (full stats)
# ----------------------------
def iter_athlete_rows(stat_group: dict) -> List[dict]:
    rows = []
    for key in ("athletes", "bench", "reserves"):
        v = stat_group.get(key)
        if isinstance(v, list):
            rows.extend(v)
    if not rows and isinstance(stat_group.get("athletes"), list):
        rows = stat_group["athletes"]
    return rows

def _idx(labels: List[str], *cands: str) -> Optional[int]:
    for c in cands:
        try:
            return labels.index(c)
        except ValueError:
            continue
    return None

def parse_player_line(values: List[str], labels: List[str]) -> Optional[dict]:
    if not labels or not values:
        return None

    i_min = _idx(labels, "MIN")
    i_fg  = _idx(labels, "FG")
    i_3pt = _idx(labels, "3PT", "3P")
    i_ft  = _idx(labels, "FT")
    i_reb = _idx(labels, "REB")
    i_ast = _idx(labels, "AST")
    i_stl = _idx(labels, "STL")
    i_blk = _idx(labels, "BLK")
    i_to  = _idx(labels, "TO")
    i_pf  = _idx(labels, "PF")
    i_pts = _idx(labels, "PTS")

    required = [i_min, i_reb, i_ast, i_stl, i_blk, i_to]
    if any(x is None for x in required):
        return None

    def get(i: Optional[int]) -> str:
        if i is None or i >= len(values):
            return ""
        return values[i]

    mins = to_minutes(get(i_min))
    fgm, fga = parse_made_attempt(get(i_fg)) if i_fg is not None else (0, 0)
    tpm, tpa = parse_made_attempt(get(i_3pt)) if i_3pt is not None else (0, 0)
    ftm, fta = parse_made_attempt(get(i_ft)) if i_ft is not None else (0, 0)

    pts = safe_int(get(i_pts)) if i_pts is not None else 0
    reb = safe_int(get(i_reb))
    ast = safe_int(get(i_ast))
    stl = safe_int(get(i_stl))
    blk = safe_int(get(i_blk))
    tov = safe_int(get(i_to))
    pf  = safe_int(get(i_pf)) if i_pf is not None else 0

    if mins == 0 and pts == 0 and reb == 0 and ast == 0 and stl == 0 and blk == 0 and tov == 0 and pf == 0 and fga == 0 and fta == 0 and tpa == 0:
        return None

    return {
        "MIN": mins,
        "PTS": pts,
        "REB": reb,
        "AST": ast,
        "STL": stl,
        "BLK": blk,
        "TO": tov,
        "PF": pf,
        "FGM": fgm, "FGA": fga,
        "3PM": tpm, "3PA": tpa,
        "FTM": ftm, "FTA": fta,
    }

def get_boxscore_players_full(event_id: str) -> List[dict]:
    url = f"{BASE}/summary?event={event_id}"
    data = get_json(url)

    box = data.get("boxscore") or {}
    players_sections = box.get("players") or []
    if not players_sections:
        return []

    out = []
    for ps in players_sections:
        team = ps.get("team", {}) or {}
        tabbr = team.get("abbreviation") or ""

        seen = set()
        for stat_group in ps.get("statistics") or []:
            labels = stat_group.get("labels") or []
            if not labels:
                continue

            for ath in iter_athlete_rows(stat_group):
                athlete = ath.get("athlete", {}) or {}
                aid = str(athlete.get("id") or "")
                pname = athlete.get("displayName") or athlete.get("shortName") or athlete.get("fullName") or "Unknown"
                values = ath.get("stats") or []

                if aid and aid in seen:
                    continue

                line = parse_player_line(values, labels)
                if not line:
                    continue

                if aid:
                    seen.add(aid)

                out.append({
                    "team": tabbr,
                    "player": pname,
                    **line
                })

    return out


# ----------------------------
# Aggregation
# ----------------------------
STAT_FIELDS_INT = ["PTS","REB","AST","STL","BLK","TO","PF","FGM","FGA","3PM","3PA","FTM","FTA"]
STAT_FIELDS_FLOAT = ["MIN"]

def agg_empty():
    return {**{k: 0 for k in STAT_FIELDS_INT}, **{k: 0.0 for k in STAT_FIELDS_FLOAT}, "G": 0}

def accumulate_player(agg: dict, line: dict):
    agg["G"] += 1
    for k in STAT_FIELDS_INT:
        agg[k] += int(line.get(k, 0))
    for k in STAT_FIELDS_FLOAT:
        agg[k] += float(line.get(k, 0.0))

def pct(made: int, att: int) -> Optional[float]:
    if att <= 0:
        return None
    return made / att

def min_per_game(agg: dict) -> str:
    g = int(agg.get("G", 0))
    if g <= 0:
        return ""
    return f"{(float(agg.get('MIN', 0.0)) / g):.1f}"


# ----------------------------
# HTML output
# ----------------------------
def write_simple_table(out_path: str, title: str, cols: List[str], rows: List[List[str]]):
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("<!doctype html><html><head><meta charset='utf-8'>")
        f.write(f"<title>{html.escape(title)}</title>")
        f.write(
            "<style>"
            "body{font-family:Arial;background:#ffffff}"
            ".wrapper{width:1100px;margin:30px auto;border:3px solid #000;background:#FFFFCC;padding:10px}"
            "h1{font-size:28px;text-align:center;background:#C0C0C0;border:1px solid #000;padding:10px;margin-top:0}"
            "table{width:100%;border-collapse:collapse;font-size:16px}"
            "th,td{border:1px solid #000;padding:6px 8px;text-align:center}"
            "th{background:#66CCFF}"
            "td.num{text-align:right}"
            "</style>"
        )
        f.write("</head><body><div class='wrapper'>")
        f.write(f"<h1>{html.escape(title)}</h1>")
        f.write("<table><thead><tr>")
        for c in cols:
            f.write(f"<th>{html.escape(c)}</th>")
        f.write("</tr></thead><tbody>")
        for r in rows:
            f.write("<tr>")
            for i, v in enumerate(r):
                cls = " class='num'" if i > 0 else ""
                f.write(f"<td{cls}>{html.escape(v)}</td>")
            f.write("</tr>")
        f.write("</tbody></table></div></body></html>")

    print(f"Wrote: {out_path}")


# ----------------------------
# Main
# ----------------------------
def main():
    cap_pd = None
    if len(sys.argv) >= 2 and sys.argv[1].strip():
        s = sys.argv[1].strip().upper()
        m = re.fullmatch(r"PD(\d+)", s)
        if not m:
            raise SystemExit("Usage: python app/build_stat_pages.py [PD7]")
        cap_pd = int(m.group(1))

    team_map = load_team_name_map()
    pd_map = load_pd_map(PD_XLSX)
    rosters = load_rosters(ROSTERS_XLSX)

    max_pd = max(pd_map.keys())
    if cap_pd is not None:
        max_pd = min(max_pd, cap_pd)

    totals_by_player = defaultdict(agg_empty)
    team_abbr_by_player: Dict[str, str] = {}
    owner_by_player: Dict[str, str] = {}  # WILL store display Team Name (mapped)

    for pd in range(1, max_pd + 1):
        if pd not in pd_map:
            continue

        primary = parse_yyyymmdd(pd_map[pd])
        prev = primary - timedelta(days=1)

        for dt in (prev, primary):
            d = fmt_yyyymmdd(dt)
            events = get_sec_events(d)

            for e in events:
                event_id = str(e.get("id") or "")
                players = get_boxscore_players_full(event_id)
                if not players:
                    continue

                for p in players:
                    key = norm_name(p.get("player", ""))
                    if not key:
                        continue

                    # roster-only
                    if key not in rosters:
                        continue

                    accumulate_player(totals_by_player[key], p)

                    if p.get("team"):
                        team_abbr_by_player[key] = p["team"]

                    old_owner = (rosters.get(key, {}) or {}).get("Team Name", "") or ""
                    owner_by_player[key] = display_team_name(old_owner, team_map)

    def roster_name(k: str) -> str:
        return rosters.get(k, {}).get("Name", "")

    def roster_field(k: str, field: str) -> str:
        return (rosters.get(k, {}) or {}).get(field, "") or ""

    # Common column block (inserted between Team and G)
    def mid_cols_values(k: str, agg: dict) -> List[str]:
        return [
            roster_field(k, "Cost"),
            roster_field(k, "Height"),
            roster_field(k, "Weight"),
            roster_field(k, "Class"),
            roster_field(k, "Pos"),
            min_per_game(agg),
        ]

    MID_COLS = ["Cost", "Height", "Weight", "Class", "Pos", "Min/G"]

    # ---------- FG% ----------
    fg_rows = []
    for k in rosters.keys():
        a = totals_by_player.get(k, agg_empty())
        fgp = pct(a["FGM"], a["FGA"])
        fg_rows.append((fgp if fgp is not None else -1.0, a["FGA"], roster_name(k), k))
    fg_rows.sort(key=lambda x: (-(x[0] if x[0] >= 0 else -1), -x[1], x[2]))

    fg_out = []
    rank = 1
    for _, __, name, k in fg_rows:
        a = totals_by_player.get(k, agg_empty())
        fgp = pct(a["FGM"], a["FGA"])
        fgp_s = f"{fgp*100:.1f}%" if fgp is not None else ""
        fg_out.append(
            [
                str(rank),
                name,
                owner_by_player.get(k, ""),
                team_abbr_by_player.get(k, roster_field(k, "Team")),
                *mid_cols_values(k, a),
                str(a["G"]) if a["G"] > 0 else "",
                f"{a['FGM']}-{a['FGA']}" if a["FGA"] > 0 else "",
                fgp_s
            ]
        )
        rank += 1

    write_simple_table(
        os.path.join(DOCS_DIR, "FieldGoalPercentage.html"),
        f"Field Goal Percentage (Through PD{max_pd})",
        ["#", "Name", "Team Name", "Team", *MID_COLS, "G", "FG", "FG%"],
        fg_out
    )

    # ---------- 3PT% ----------
    tp_rows = []
    for k in rosters.keys():
        a = totals_by_player.get(k, agg_empty())
        p3 = pct(a["3PM"], a["3PA"])
        tp_rows.append((p3 if p3 is not None else -1.0, a["3PA"], roster_name(k), k))
    tp_rows.sort(key=lambda x: (-(x[0] if x[0] >= 0 else -1), -x[1], x[2]))

    tp_out = []
    rank = 1
    for _, __, name, k in tp_rows:
        a = totals_by_player.get(k, agg_empty())
        p3 = pct(a["3PM"], a["3PA"])
        p3_s = f"{p3*100:.1f}%" if p3 is not None else ""
        tp_out.append(
            [
                str(rank),
                name,
                owner_by_player.get(k, ""),
                team_abbr_by_player.get(k, roster_field(k, "Team")),
                *mid_cols_values(k, a),
                str(a["G"]) if a["G"] > 0 else "",
                f"{a['3PM']}-{a['3PA']}" if a["3PA"] > 0 else "",
                p3_s
            ]
        )
        rank += 1

    write_simple_table(
        os.path.join(DOCS_DIR, "ThreePointFieldGoalPercentage.html"),
        f"3-Point Field Goal Percentage (Through PD{max_pd})",
        ["#", "Name", "Team Name", "Team", *MID_COLS, "G", "3PT", "3PT%"],
        tp_out
    )

    # ---------- FT% ----------
    ft_rows = []
    for k in rosters.keys():
        a = totals_by_player.get(k, agg_empty())
        ftp = pct(a["FTM"], a["FTA"])
        ft_rows.append((ftp if ftp is not None else -1.0, a["FTA"], roster_name(k), k))
    ft_rows.sort(key=lambda x: (-(x[0] if x[0] >= 0 else -1), -x[1], x[2]))

    ft_out = []
    rank = 1
    for _, __, name, k in ft_rows:
        a = totals_by_player.get(k, agg_empty())
        ftp = pct(a["FTM"], a["FTA"])
        ftp_s = f"{ftp*100:.1f}%" if ftp is not None else ""
        ft_out.append(
            [
                str(rank),
                name,
                owner_by_player.get(k, ""),
                team_abbr_by_player.get(k, roster_field(k, "Team")),
                *mid_cols_values(k, a),
                str(a["G"]) if a["G"] > 0 else "",
                f"{a['FTM']}-{a['FTA']}" if a["FTA"] > 0 else "",
                ftp_s
            ]
        )
        rank += 1

    write_simple_table(
        os.path.join(DOCS_DIR, "FreeThrowPercentage.html"),
        f"Free Throw Percentage (Through PD{max_pd})",
        ["#", "Name", "Team Name", "Team", *MID_COLS, "G", "FT", "FT%"],
        ft_out
    )

    # Count-stat per game pages
    def write_count_page(filename: str, label: str, field: str):
        rows = []
        for k in rosters.keys():
            a = totals_by_player.get(k, agg_empty())
            g = a["G"]
            total = a[field]
            per_g = (total / g) if g > 0 else None
            rows.append((per_g if per_g is not None else -1.0, total, roster_name(k), k))

        rows.sort(key=lambda x: (-(x[0] if x[0] >= 0 else -1), -x[1], x[2]))

        out = []
        rank = 1
        for _, __, name, k in rows:
            a = totals_by_player.get(k, agg_empty())
            g = a["G"]
            total = a[field]
            per_g = (total / g) if g > 0 else None
            out.append(
                [
                    str(rank),
                    name,
                    owner_by_player.get(k, ""),
                    team_abbr_by_player.get(k, roster_field(k, "Team")),
                    *mid_cols_values(k, a),
                    str(g) if g > 0 else "",
                    str(total) if g > 0 else "",
                    f"{per_g:.2f}" if per_g is not None else ""
                ]
            )
            rank += 1

        write_simple_table(
            os.path.join(DOCS_DIR, filename),
            f"{label} (Per Game) (Through PD{max_pd})",
            ["#", "Name", "Team Name", "Team", *MID_COLS, "G", f"{label} (Total)", f"{label}/G"],
            out
        )

    write_count_page("Rebounds.html",      "Rebounds",       "REB")
    write_count_page("BlockedShots.html",  "Blocked Shots",  "BLK")
    write_count_page("Assists.html",       "Assists",        "AST")
    write_count_page("Steals.html",        "Steals",         "STL")
    write_count_page("Turnovers.html",     "Turnovers",      "TO")
    write_count_page("PersonalFouls.html", "Personal Fouls", "PF")


if __name__ == "__main__":
    main()
